#!/usr/bin/env python3
"""
One-shot analysis: cross-account shared IPs from existing login JSON artifacts.

Reads the `analysis_ip_to_accounts.json` files already produced by the
Login IP Monitor's daily pipeline and reports, for the last N days:

  1. 单日视图 (per-day)    : 每一天内被 ≥2 个账户登录的 IP 数量 / 涉及账户数
  2. 窗口累计视图 (window) : N 天内累计出现过 ≥2 个账户的 IP（跨天合并）
  3. 分布 & Top-K 列表     : IP 被几个账户共享的分布 + 最"乱"的前 K 个 IP

Read-only：**不写 DB、不改文件**。仅 stdout（以及可选 CSV）。

为什么用现有 JSON 而不是重新解析 .log
----------------------------------------
JSON 里 `{server: {ip: [account_ids]}}` 已经是全量登录账户（不限 watchlist）。
Parser pass-1 会把所有 `': login` 事件都累加到 `ip_to_accounts`，所以我们
不用碰 GB 级原始日志就能做跨账户 IP 聚合。

Usage
-----
    cd /opt/myproject/New-IT-System/backend
    source .venv/bin/activate

    # 默认：截至昨天的过去 14 天
    python scripts/analyze_shared_ip_cross_account.py

    # 指定区间
    python scripts/analyze_shared_ip_cross_account.py --start 20260410 --end 20260423

    # 指定天数（从今天往前数 N 天，含今天）
    python scripts/analyze_shared_ip_cross_account.py --days 7

    # 导出 CSV（所有 IP-日-账户的 flat 行）
    python scripts/analyze_shared_ip_cross_account.py --csv /tmp/shared_ip.csv

    # 排除服务器内部账号（account_id < 阈值，默认 1000）
    python scripts/analyze_shared_ip_cross_account.py --min-account-id 1000

    # 输出每个服务器独立统计（默认是合并后的 cross-server 视角）
    python scripts/analyze_shared_ip_cross_account.py --per-server
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

BACKEND_ROOT = Path(__file__).resolve().parents[1]
# Make `from app.services...` importable when run as a plain script (not -m).
sys.path.insert(0, str(BACKEND_ROOT))

DATA_DIR = BACKEND_ROOT / "data" / "login_ip"
JSON_FILENAME = "analysis_ip_to_accounts.json"

# Maps the server_name used inside analysis JSON to the numeric `sid` prefix
# stored in `fxbackoffice.mt4_trades.loginSid` (format: `{sid}-{LOGIN}`).
# Matches `mt4_users.sid` usage in client_return_service.py (`sid IN (1, 5, 6)`).
SERVER_TO_SID_PREFIX: dict[str, str] = {
    "MT4": "1-",
    "MT5": "5-",
    "MT4_Live2": "6-",
}

# IN-list chunk size for the trade-count lookup. 1000 is well under
# max_allowed_packet and matches login_ip_enrichment_service._IN_CLAUSE_CHUNK.
_TRADE_IN_CLAUSE_CHUNK = 1000


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------


def parse_date_arg(s: str) -> dt.date:
    """Parse YYYYMMDD."""
    return dt.datetime.strptime(s, "%Y%m%d").date()


def resolve_date_range(args: argparse.Namespace) -> list[dt.date]:
    """Decide which dates to analyze based on --date / --start/--end / --days."""
    if args.date:
        # Single-day shortcut — most common ad-hoc review usage.
        one = parse_date_arg(args.date)
        return [one]
    if args.start and args.end:
        start = parse_date_arg(args.start)
        end = parse_date_arg(args.end)
    else:
        # --days N means: the last N days ending yesterday (same convention
        # as backfill_login_ip.py, so "yesterday" is always in-window).
        days = args.days
        end = dt.date.today() - dt.timedelta(days=1)
        start = end - dt.timedelta(days=days - 1)

    if start > end:
        sys.exit(f"start ({start}) must be <= end ({end})")

    # Inclusive range
    return [start + dt.timedelta(days=i) for i in range((end - start).days + 1)]


def date_str(d: dt.date) -> str:
    return d.strftime("%Y%m%d")


# ---------------------------------------------------------------------------
# JSON loading
# ---------------------------------------------------------------------------


def load_day(date: dt.date) -> dict[str, dict[str, list[int]]] | None:
    """Return {server: {ip: [account_ids]}} for one day, or None if missing."""
    path = DATA_DIR / date_str(date) / JSON_FILENAME
    if not path.exists():
        return None
    with open(path, "r", encoding="utf-8") as fp:
        return json.load(fp)


def filter_accounts(
    accounts: Iterable[int],
    min_account_id: int,
    known_accounts: set[int] | None = None,
) -> list[int]:
    """Drop server-internal accounts + (optional) accounts not mapped in CRM.

    - min_account_id: floor out accounts like 25 / 776 (MT server-internal).
    - known_accounts: if provided, only keep accounts that exist in CRM
      `fxbackoffice.mt4_users` with a non-null `userId`, i.e. real clients.
    """
    out: set[int] = set()
    for a in accounts:
        ai = int(a)
        if ai < min_account_id:
            continue
        if known_accounts is not None and ai not in known_accounts:
            continue
        out.add(ai)
    return sorted(out)


def load_known_accounts_from_crm() -> tuple[set[int], dict[int, dict]]:
    """Pull the whitelist of MT LOGINs that have a real CRM user attached.

    Returns `(login_set, {login: {client_id, chinese_name}})`. Uses the same
    MySQL slave + credentials as the Login IP enrichment service. Runs ONE
    SELECT per call (no IN-chunking — we want the FULL list, not a subset).
    """
    # Import lazily so the script still works when CRM is unreachable AND
    # --known-accounts-only is not used.
    import pymysql
    import pymysql.cursors
    from app.core.config import get_settings

    settings = get_settings()
    conn = pymysql.connect(
        host=settings.DB_HOST,
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        database=settings.MYSQL_DATABASE_FXBACKOFFICE,
        port=int(settings.DB_PORT),
        charset=settings.DB_CHARSET,
        cursorclass=pymysql.cursors.DictCursor,
    )
    login_set: set[int] = set()
    details: dict[int, dict] = {}
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT u.LOGIN AS login, u.userId AS client_id, cf.v AS chinese_name
                FROM fxbackoffice.mt4_users u
                LEFT JOIN fxbackoffice.user_custom_fields cf
                  ON u.userId = cf.userid AND cf.k = 'custom_chinese_name'
                WHERE u.userId IS NOT NULL
                """
            )
            for row in cursor.fetchall():
                try:
                    login = int(row["login"])
                except (TypeError, ValueError):
                    continue
                # Upstream `user_custom_fields.v` wraps values in literal
                # double-quotes (e.g. `"张三"`). Strip to keep CSV clean.
                cname = row.get("chinese_name") or ""
                if isinstance(cname, str):
                    cname = cname.strip().strip('"').strip()
                login_set.add(login)
                details[login] = {
                    "client_id": row["client_id"],
                    "chinese_name": cname or None,
                }
    finally:
        conn.close()
    return login_set, details


def load_trade_counts_from_crm(
    login_sids: Iterable[str], trade_date: dt.date
) -> dict[str, int]:
    """Return `{loginSid: order_count}` for one date.

    Only real market fills (CMD 0/1) and non-deleted rows are counted. Uses
    the `openDate` STORED generated column (has its own `IDX_OPEN_DATE` index)
    so the query is bounded to a single day's worth of rows before the
    loginSid IN-filter is applied — MySQL optimizer typically picks the date
    index first, which is exactly what we want here.

    loginSids NOT present in the returned dict have zero matching orders.
    """
    import pymysql
    import pymysql.cursors
    from app.core.config import get_settings

    ids = [s for s in login_sids if s]
    if not ids:
        return {}

    settings = get_settings()
    # Explicit timeouts so Azure slave hiccups surface as errors within ~30s
    # instead of hanging forever. read_timeout 60s covers the ~1-3s query +
    # network jitter on bigger chunks.
    t_conn = dt.datetime.now()
    print(f"  [sql] connecting to {settings.DB_HOST} ...", flush=True)
    conn = pymysql.connect(
        host=settings.DB_HOST,
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        database=settings.MYSQL_DATABASE_FXBACKOFFICE,
        port=int(settings.DB_PORT),
        charset=settings.DB_CHARSET,
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=10,
        read_timeout=60,
        write_timeout=30,
    )
    print(
        f"  [sql] connected in {(dt.datetime.now() - t_conn).total_seconds():.2f}s",
        flush=True,
    )
    counts: dict[str, int] = {}
    date_str_sql = trade_date.strftime("%Y-%m-%d")
    total_chunks = (len(ids) + _TRADE_IN_CLAUSE_CHUNK - 1) // _TRADE_IN_CLAUSE_CHUNK
    try:
        with conn.cursor() as cursor:
            # Chunk the IN list for safety even though 1197 × ~10 char = 12KB
            # easily fits; chunking is cheap and protects future runs
            # with 14-day windows that may have 10k+ distinct loginSids.
            for chunk_idx, start in enumerate(
                range(0, len(ids), _TRADE_IN_CLAUSE_CHUNK), start=1
            ):
                chunk = ids[start : start + _TRADE_IN_CLAUSE_CHUNK]
                placeholders = ", ".join(["%s"] * len(chunk))
                # Use sargable `(isDeleted = 0 OR IS NULL)` — COALESCE wraps
                # the column in a function and defeats index selection.
                # EXPLAIN confirms planner picks IDX_OPEN_DATE with this shape.
                sql = f"""
                    SELECT loginSid, COUNT(*) AS order_count
                    FROM fxbackoffice.mt4_trades
                    WHERE openDate = %s
                      AND CMD IN (0, 1)
                      AND (isDeleted = 0 OR isDeleted IS NULL)
                      AND loginSid IN ({placeholders})
                    GROUP BY loginSid
                """
                t0 = dt.datetime.now()
                print(
                    f"  [sql] chunk {chunk_idx}/{total_chunks} "
                    f"({len(chunk)} loginSids) ...",
                    flush=True,
                )
                cursor.execute(sql, (date_str_sql, *chunk))
                rows_fetched = 0
                for row in cursor.fetchall():
                    counts[str(row["loginSid"])] = int(row["order_count"])
                    rows_fetched += 1
                print(
                    f"  [sql] chunk {chunk_idx} done in "
                    f"{(dt.datetime.now() - t0).total_seconds():.2f}s "
                    f"→ {rows_fetched} loginSids with trades",
                    flush=True,
                )
    finally:
        conn.close()
    return counts


# ---------------------------------------------------------------------------
# Core aggregation
# ---------------------------------------------------------------------------


def aggregate(
    dates: list[dt.date],
    min_account_id: int,
    per_server: bool,
    known_accounts: set[int] | None = None,
) -> tuple[dict, dict, list[str], dict[int, set[str]]]:
    """
    Returns (per_day_stats, window_stats, missing_dates, account_servers).

    per_day_stats   : {date_str: {server_or_all: {...metrics...}}}
    window_stats    : {server_or_all: {ip: {accounts: set[int], days: set[str]}}}
    account_servers : {account_id: {"MT4", "MT5", ...}} — needed to rebuild the
      `loginSid = {sid}-{LOGIN}` key for the trade-enrichment lookup regardless
      of whether we aggregated per-server or merged.
    """
    # For the window view we need to know each IP's full account set across
    # the whole period, plus which days it was seen on (so we can answer
    # "new shared IPs this week vs last week" later if needed).
    window: dict[str, dict[str, dict]] = defaultdict(
        lambda: defaultdict(lambda: {"accounts": set(), "days": set()})
    )
    per_day: dict[str, dict] = {}
    missing: list[str] = []
    account_servers: dict[int, set[str]] = defaultdict(set)

    for d in dates:
        ds = date_str(d)
        day_json = load_day(d)
        if day_json is None:
            missing.append(ds)
            continue

        # For each server, or merged if not per_server
        day_metrics: dict[str, dict] = {}

        # Temporary per-day aggregation buckets
        day_ip_to_accs: dict[str, dict[str, set[int]]] = defaultdict(
            lambda: defaultdict(set)
        )

        for server, ip_map in day_json.items():
            bucket_key = server if per_server else "ALL"
            for ip, accs in ip_map.items():
                cleaned = filter_accounts(accs, min_account_id, known_accounts)
                if not cleaned:
                    continue
                day_ip_to_accs[bucket_key][ip].update(cleaned)
                # Window-level aggregation
                window[bucket_key][ip]["accounts"].update(cleaned)
                window[bucket_key][ip]["days"].add(ds)
                # Track which server(s) each account came from. An account ID
                # usually lives on exactly one server, but theoretically MT5
                # and MT4 can share an LOGIN — storing a set protects us.
                for a in cleaned:
                    account_servers[a].add(server)

        # Build day metrics
        for bucket_key, ip_map in day_ip_to_accs.items():
            shared_ips = {ip: accs for ip, accs in ip_map.items() if len(accs) >= 2}
            accounts_involved: set[int] = set()
            for accs in shared_ips.values():
                accounts_involved.update(accs)

            day_metrics[bucket_key] = {
                "total_ips": len(ip_map),
                "shared_ips": len(shared_ips),
                "accounts_involved": len(accounts_involved),
                "shared_ip_accounts_distribution": Counter(
                    len(accs) for accs in shared_ips.values()
                ),
            }

        per_day[ds] = day_metrics

    return per_day, window, missing, dict(account_servers)


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------


def _bucket(n: int) -> str:
    """Group 'accounts per IP' into readable buckets."""
    if n == 2:
        return "2"
    if n == 3:
        return "3"
    if 4 <= n <= 5:
        return "4-5"
    if 6 <= n <= 10:
        return "6-10"
    if 11 <= n <= 20:
        return "11-20"
    return ">20"


BUCKET_ORDER = ["2", "3", "4-5", "6-10", "11-20", ">20"]


def print_per_day(per_day: dict) -> None:
    print("=" * 78)
    print("单日视图 (per-day): 当天同一 IP 登录账户数 ≥2 的情况")
    print("=" * 78)
    header = f"{'Date':10}  {'Server':12}  {'TotalIPs':>9}  {'SharedIPs':>10}  {'Accounts':>9}"
    print(header)
    print("-" * len(header))
    for ds in sorted(per_day.keys()):
        for bucket_key, m in sorted(per_day[ds].items()):
            print(
                f"{ds:10}  {bucket_key:12}  {m['total_ips']:>9,}  "
                f"{m['shared_ips']:>10,}  {m['accounts_involved']:>9,}"
            )
    print()


def print_window(window: dict, window_days: int) -> None:
    print("=" * 78)
    print(f"窗口累计视图 ({window_days} 天): 期内被 ≥2 个账户用过的 IP")
    print("=" * 78)
    for bucket_key in sorted(window.keys()):
        ip_map = window[bucket_key]
        shared = {ip: v for ip, v in ip_map.items() if len(v["accounts"]) >= 2}
        accounts_all: set[int] = set()
        for v in shared.values():
            accounts_all.update(v["accounts"])

        print(f"\n[{bucket_key}]")
        print(f"  unique IPs seen       : {len(ip_map):,}")
        print(f"  shared IPs (≥2 accts) : {len(shared):,}")
        print(f"  distinct accts touched: {len(accounts_all):,}")

        # Distribution
        dist = Counter(_bucket(len(v["accounts"])) for v in shared.values())
        print("  accounts-per-IP distribution:")
        for b in BUCKET_ORDER:
            if dist.get(b):
                print(f"    {b:>5} accts : {dist[b]:,} IPs")


def print_top_k(window: dict, k: int, min_account_id: int) -> None:
    print()
    print("=" * 78)
    print(f"Top {k} 最多账户共享的 IP（窗口累计）")
    print("=" * 78)
    for bucket_key in sorted(window.keys()):
        ip_map = window[bucket_key]
        ranked = sorted(
            (
                (ip, v["accounts"], v["days"])
                for ip, v in ip_map.items()
                if len(v["accounts"]) >= 2
            ),
            key=lambda x: (-len(x[1]), -len(x[2]), x[0]),
        )
        print(f"\n[{bucket_key}]")
        for ip, accs, days in ranked[:k]:
            days_str = ", ".join(sorted(days))
            acc_preview = ", ".join(str(a) for a in sorted(accs)[:8])
            if len(accs) > 8:
                acc_preview += f", ... (+{len(accs) - 8} more)"
            print(
                f"  {ip:<20}  accts={len(accs):>3}  days={len(days):>2}  "
                f"[{acc_preview}]"
            )
            print(f"    └─ days: {days_str}")


def _distinct_client_ids(
    accs: Iterable[int], details: dict[int, dict] | None
) -> int | str:
    """Count unique CRM UIDs across the given accounts. '-' if details missing."""
    if not details:
        return "-"
    uids = {details[a]["client_id"] for a in accs if a in details and details[a].get("client_id") is not None}
    return len(uids)


def _flatten_rows(
    window: dict,
    details: dict[int, dict] | None,
    trade_counts: dict[int, int] | None = None,
) -> list[dict]:
    """Build a single sorted list of per-account rows across all buckets.

    Sort order makes it read top-to-bottom as "most suspicious first":
        1. distinct_client_ids DESC   (more real customers sharing 1 IP = worse)
        2. accounts_on_ip DESC
        3. ip ASC (stable grouping)
        4. client_id ASC within the same IP (same client's accounts cluster)

    `trade_counts` (optional) is `{account_id: orders_on_date}` from
    load_trade_counts_from_crm(); missing keys mean 0 orders.
    """
    rows: list[dict] = []
    for bucket_key, ip_map in window.items():
        for ip, v in ip_map.items():
            if len(v["accounts"]) < 2:
                continue
            accs_sorted = sorted(v["accounts"])
            n_accs = len(accs_sorted)
            n_days = len(v["days"])
            days_str = "|".join(sorted(v["days"]))
            distinct_uids_raw = _distinct_client_ids(accs_sorted, details)
            distinct_uids = distinct_uids_raw if isinstance(distinct_uids_raw, int) else 0
            for acc in accs_sorted:
                info = (details or {}).get(acc, {}) or {}
                orders = (trade_counts or {}).get(acc)
                rows.append({
                    "ip": ip,
                    "client_id": info.get("client_id") or "",
                    "chinese_name": info.get("chinese_name") or "",
                    "mt_account": acc,
                    "server": bucket_key,
                    "clients_on_ip": distinct_uids_raw,
                    "accounts_on_ip": n_accs,
                    "days_on_ip": n_days,
                    "days": days_str,
                    # Only populated when --enrich-trades was used; otherwise
                    # left as None so the writers can skip the columns entirely.
                    "orders_on_date": orders if trade_counts is not None else None,
                    "traded_on_date": (
                        int(bool(orders)) if trade_counts is not None else None
                    ),
                    "_sort_uids": distinct_uids,
                    "_sort_client_id": info.get("client_id") or 0,
                })
    rows.sort(
        key=lambda r: (
            -r["_sort_uids"],
            -r["accounts_on_ip"],
            r["ip"],
            r["_sort_client_id"] if isinstance(r["_sort_client_id"], int) else 0,
            r["mt_account"],
        )
    )
    for r in rows:
        r.pop("_sort_uids", None)
        r.pop("_sort_client_id", None)
    return rows


def write_csv(
    window: dict,
    csv_path: Path,
    details: dict[int, dict] | None,
    trade_counts: dict[int, int] | None = None,
    trade_date: dt.date | None = None,
) -> None:
    """One row per (ip, account), grouped by IP, sorted by suspicion DESC.

    When `trade_counts` is provided, two extra columns are appended:
      - `orders_<YYYYMMDD>`  : number of fills (CMD 0/1) on that date.
      - `traded_<YYYYMMDD>`  : 1 if orders > 0 else 0, for quick filtering.
    """
    rows = _flatten_rows(window, details, trade_counts)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    date_suffix = date_str(trade_date) if trade_date else "date"
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.writer(fp)
        header = [
            "ip", "client_id", "chinese_name", "mt_account", "server",
            "clients_on_ip", "accounts_on_ip", "days_on_ip", "days",
        ]
        if trade_counts is not None:
            header += [f"orders_{date_suffix}", f"traded_{date_suffix}"]
        writer.writerow(header)
        for r in rows:
            row = [
                r["ip"], r["client_id"], r["chinese_name"], r["mt_account"],
                r["server"], r["clients_on_ip"], r["accounts_on_ip"],
                r["days_on_ip"], r["days"],
            ]
            if trade_counts is not None:
                row += [r.get("orders_on_date") or 0, r.get("traded_on_date") or 0]
            writer.writerow(row)
    print(f"\nCSV written: {csv_path}  ({len(rows):,} rows)")


def write_xlsx(
    window: dict,
    xlsx_path: Path,
    details: dict[int, dict] | None,
    trade_counts: dict[int, int] | None = None,
    trade_date: dt.date | None = None,
) -> None:
    """Produce a human-friendly Excel workbook with 2 sheets.

    Sheet 1 「按 IP 分组」: one row per (ip, client) — IP cell is merged across
      its block, rows are color-banded per IP for scannability, and the
      `clients_on_ip` column has a red/orange/yellow gradient. When
      `trade_counts` is supplied, appends "当日订单数" (red if >0 so it jumps
      out) and "当日交易?" columns.
    Sheet 2 「IP 汇总」    : one row per IP for quick overview / sorting. When
      `trade_counts` is supplied, adds "当日活跃客户数/账户数" columns — high
      values = many real traders behind that IP = highest priority to review.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter

    rows = _flatten_rows(window, details, trade_counts)
    date_label = trade_date.strftime("%m/%d") if trade_date else "当日"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "按 IP 分组"

    # ------------------------------------------------------------------
    # Style palette
    # ------------------------------------------------------------------
    header_fill = PatternFill("solid", fgColor="1F2937")  # slate-800
    header_font = Font(bold=True, color="FFFFFF", size=11)
    band_fills = [
        PatternFill("solid", fgColor="FFFFFF"),
        PatternFill("solid", fgColor="F1F5F9"),  # slate-100
    ]
    # Severity colors for `clients_on_ip`.
    sev_red = PatternFill("solid", fgColor="FCA5A5")     # >= 10 UIDs
    sev_orange = PatternFill("solid", fgColor="FDBA74")  # 5-9 UIDs
    sev_yellow = PatternFill("solid", fgColor="FEF08A")  # 3-4 UIDs
    thin = Side(style="thin", color="CBD5E1")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Highlight fills for trade-activity columns (only used when
    # --enrich-trades was passed).
    trade_yes_fill = PatternFill("solid", fgColor="FECACA")  # red-200
    trade_no_fill = PatternFill("solid", fgColor="DCFCE7")   # green-100

    # ------------------------------------------------------------------
    # Sheet 1: per-account rows, IP merged vertically
    # ------------------------------------------------------------------
    headers1 = [
        "IP", "Client ID", "中文名", "MT 账户", "服务器",
        "客户数 (UIDs)", "账户数", "出现天数",
    ]
    if trade_counts is not None:
        headers1 += [f"{date_label} 订单数", f"{date_label} 交易?"]
    ws1.append(headers1)
    for col_idx in range(1, len(headers1) + 1):
        c = ws1.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = cell_border

    def _severity_fill(uids: int) -> PatternFill | None:
        if uids >= 10:
            return sev_red
        if uids >= 5:
            return sev_orange
        if uids >= 3:
            return sev_yellow
        return None

    # Walk in stable order; start row after header.
    row_cursor = 2
    band_idx = 0
    prev_ip: str | None = None
    ip_block_start = row_cursor
    for r in rows:
        ip = r["ip"]
        if ip != prev_ip:
            if prev_ip is not None and row_cursor - 1 > ip_block_start:
                ws1.merge_cells(
                    start_row=ip_block_start, end_row=row_cursor - 1,
                    start_column=1, end_column=1,
                )
            prev_ip = ip
            ip_block_start = row_cursor
            band_idx = 1 - band_idx  # alternate band per new IP

        band_fill = band_fills[band_idx]
        uids = r["clients_on_ip"] if isinstance(r["clients_on_ip"], int) else 0
        sev = _severity_fill(uids)

        values = [
            ip, r["client_id"], r["chinese_name"], r["mt_account"], r["server"],
            r["clients_on_ip"], r["accounts_on_ip"], r["days_on_ip"],
        ]
        if trade_counts is not None:
            # None → not looked up; 0 → looked up, no trades. We still render
            # 0 as "0" so readers can visually confirm coverage.
            orders_val = r.get("orders_on_date") or 0
            traded_val = "是" if orders_val else "否"
            values += [orders_val, traded_val]
        for col_idx, val in enumerate(values, start=1):
            c = ws1.cell(row=row_cursor, column=col_idx, value=val)
            c.border = cell_border
            c.alignment = center if col_idx != 3 else left_wrap
            # Severity override on `clients_on_ip` column (col 6), else band.
            if col_idx == 6 and sev is not None:
                c.fill = sev
            elif trade_counts is not None and col_idx in (9, 10):
                # Trade columns get red/green highlight to pop.
                c.fill = trade_yes_fill if (r.get("orders_on_date") or 0) > 0 else trade_no_fill
                if col_idx == 9 and (r.get("orders_on_date") or 0) > 0:
                    c.font = Font(bold=True, color="991B1B")
            else:
                c.fill = band_fill
        row_cursor += 1

    # Close the final IP block merge.
    if prev_ip is not None and row_cursor - 1 > ip_block_start:
        ws1.merge_cells(
            start_row=ip_block_start, end_row=row_cursor - 1,
            start_column=1, end_column=1,
        )

    ws1.freeze_panes = "A2"
    col_widths1 = [18, 12, 14, 14, 12, 14, 10, 12]
    if trade_counts is not None:
        col_widths1 += [14, 11]
    for i, w in enumerate(col_widths1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # Sheet 2: per-IP summary
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("IP 汇总")
    headers2 = [
        "IP", "服务器", "客户数 (UIDs)", "账户数", "出现天数",
        "首日", "末日", "账户列表", "Client IDs", "中文名",
    ]
    if trade_counts is not None:
        # Added at the end so existing columns keep their positions.
        headers2 += [f"{date_label} 活跃账户", f"{date_label} 总订单"]
    ws2.append(headers2)
    for col_idx in range(1, len(headers2) + 1):
        c = ws2.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = cell_border

    summary_rows = []
    for bucket_key, ip_map in window.items():
        for ip, v in ip_map.items():
            if len(v["accounts"]) < 2:
                continue
            accs_sorted = sorted(v["accounts"])
            days_sorted = sorted(v["days"])
            uids = _distinct_client_ids(accs_sorted, details)
            client_ids_str = ""
            names_str = ""
            if details:
                client_ids_str = "|".join(
                    str((details.get(a, {}) or {}).get("client_id") or "")
                    for a in accs_sorted
                )
                names_str = "|".join(
                    ((details.get(a, {}) or {}).get("chinese_name") or "")
                    for a in accs_sorted
                )
            row_summary = [
                ip, bucket_key, uids, len(accs_sorted), len(days_sorted),
                days_sorted[0], days_sorted[-1],
                "|".join(str(a) for a in accs_sorted),
                client_ids_str, names_str,
            ]
            if trade_counts is not None:
                active = sum(1 for a in accs_sorted if (trade_counts.get(a) or 0) > 0)
                total_orders = sum(trade_counts.get(a, 0) or 0 for a in accs_sorted)
                row_summary += [active, total_orders]
            summary_rows.append(row_summary)
    # Sort: if trade enrichment is on, prioritize IPs whose shared accounts
    # ACTUALLY traded — those are the real risk signals — then fall back to
    # UIDs/accounts/days. Otherwise use the old UIDs-first order.
    def _summary_sort_key(r: list) -> tuple:
        active = r[10] if trade_counts is not None and len(r) > 10 else 0
        uids_val = r[2] if isinstance(r[2], int) else 0
        return (-active, -uids_val, -r[3], -r[4], r[0])
    summary_rows.sort(key=_summary_sort_key)

    row_cursor = 2
    for i, r in enumerate(summary_rows):
        band_fill = band_fills[i % 2]
        uids = r[2] if isinstance(r[2], int) else 0
        sev = _severity_fill(uids)
        active = r[10] if trade_counts is not None and len(r) > 10 else 0
        for col_idx, val in enumerate(r, start=1):
            c = ws2.cell(row=row_cursor, column=col_idx, value=val)
            c.border = cell_border
            c.alignment = center if col_idx not in (8, 9, 10) else left_wrap
            if col_idx == 3 and sev is not None:
                c.fill = sev
            elif trade_counts is not None and col_idx in (11, 12):
                # Highlight IPs with any real trading activity in red.
                c.fill = trade_yes_fill if active > 0 else trade_no_fill
                if active > 0:
                    c.font = Font(bold=True, color="991B1B")
            else:
                c.fill = band_fill
        row_cursor += 1

    ws2.freeze_panes = "A2"
    col_widths2 = [18, 12, 14, 10, 10, 12, 12, 60, 50, 40]
    if trade_counts is not None:
        col_widths2 += [14, 14]
    for i, w in enumerate(col_widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(
        f"XLSX written: {xlsx_path}  "
        f"(sheet1 rows={len(rows):,}, sheet2 rows={len(summary_rows):,})"
    )


def write_csv_pivot(
    window: dict,
    csv_path: Path,
    details: dict[int, dict] | None,
    trade_counts: dict[int, int] | None = None,
    trade_date: dt.date | None = None,
) -> None:
    """One row per IP (accounts pipe-joined) — human-friendly overview."""
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    date_suffix = date_str(trade_date) if trade_date else "date"
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.writer(fp)
        header = [
            "server_bucket", "ip", "accounts_count", "distinct_client_ids",
            "days_count", "first_day", "last_day",
            "accounts", "client_ids", "chinese_names", "days",
        ]
        if trade_counts is not None:
            header += [
                f"active_accounts_{date_suffix}",
                f"total_orders_{date_suffix}",
            ]
        writer.writerow(header)
        rows = []
        for bucket_key, ip_map in window.items():
            for ip, v in ip_map.items():
                if len(v["accounts"]) < 2:
                    continue
                days_sorted = sorted(v["days"])
                accs_sorted = sorted(v["accounts"])
                distinct_uids = _distinct_client_ids(accs_sorted, details)
                client_ids_str = ""
                names_str = ""
                if details:
                    client_ids_str = "|".join(
                        str(details.get(a, {}).get("client_id", "") or "")
                        for a in accs_sorted
                    )
                    names_str = "|".join(
                        (details.get(a, {}).get("chinese_name", "") or "")
                        for a in accs_sorted
                    )
                row_tuple = [
                    bucket_key, ip, len(accs_sorted), distinct_uids,
                    len(days_sorted), days_sorted[0], days_sorted[-1],
                    "|".join(str(a) for a in accs_sorted),
                    client_ids_str, names_str,
                    "|".join(days_sorted),
                ]
                if trade_counts is not None:
                    active = sum(1 for a in accs_sorted if (trade_counts.get(a) or 0) > 0)
                    total_orders = sum(trade_counts.get(a, 0) or 0 for a in accs_sorted)
                    row_tuple += [active, total_orders]
                rows.append(row_tuple)
        # Sort: with trades → real activity first; else keep UIDs-first.
        def _sort_key(r):
            uids = r[3] if isinstance(r[3], int) else -1
            active = r[11] if trade_counts is not None and len(r) > 11 else 0
            return (-active, -uids, -r[2], -r[4], r[1])
        rows.sort(key=_sort_key)
        for row in rows:
            writer.writerow(row)
    print(f"CSV (pivot) written: {csv_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Cross-account shared-IP analysis over existing login JSONs.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--date", help="YYYYMMDD single-day shortcut (overrides other range flags)")
    parser.add_argument("--start", help="YYYYMMDD (inclusive)")
    parser.add_argument("--end", help="YYYYMMDD (inclusive)")
    parser.add_argument(
        "--days",
        type=int,
        default=14,
        help="Last N days ending yesterday if --date / --start/--end not provided (default 14)",
    )
    parser.add_argument(
        "--min-account-id",
        type=int,
        default=1000,
        help="Drop account IDs below this (server-internal manager accounts). Default 1000",
    )
    parser.add_argument(
        "--known-accounts-only",
        action="store_true",
        help="Only keep MT accounts that exist in CRM fxbackoffice.mt4_users with userId NOT NULL. "
             "Also enriches CSVs with client_id + chinese_name. Needs DB_* env vars.",
    )
    parser.add_argument(
        "--per-server",
        action="store_true",
        help="Break down per MT4/MT5/MT4_Live2. Default: merge across servers.",
    )
    parser.add_argument(
        "--top-k", type=int, default=15, help="Print top-K noisiest IPs (default 15)"
    )
    parser.add_argument("--csv", help="Optional: flat CSV (one row per IP × account, sorted by suspicion)")
    parser.add_argument(
        "--csv-pivot",
        help="Optional: pivoted CSV (one row per IP, accounts pipe-joined)",
    )
    parser.add_argument(
        "--xlsx",
        help="Optional: Excel workbook (2 sheets: per-account + per-IP summary, "
             "with merged cells / frozen header / severity color-coding)",
    )
    parser.add_argument(
        "--enrich-trades",
        metavar="YYYYMMDD",
        help="Optional: look up how many fills each shared-IP account made on "
             "this date in fxbackoffice.mt4_trades (CMD 0/1, not deleted). "
             "Adds `orders_<date>` + `traded_<date>` columns to every output. "
             "Requires DB_* env vars (same as --known-accounts-only).",
    )
    args = parser.parse_args()

    dates = resolve_date_range(args)
    print(
        f"Analyzing {len(dates)} day(s): {date_str(dates[0])} → {date_str(dates[-1])}"
    )
    print(
        f"min_account_id = {args.min_account_id}   "
        f"per_server = {args.per_server}   "
        f"known_accounts_only = {args.known_accounts_only}"
    )

    # Load .env once if CRM lookups will be needed for any reason (known-accounts
    # filter OR trade-enrichment). Safer than reloading per branch.
    needs_db = args.known_accounts_only or bool(args.enrich_trades)
    if needs_db:
        try:
            from dotenv import load_dotenv
            load_dotenv(BACKEND_ROOT / ".env")
        except ImportError:
            pass

    known_accounts: set[int] | None = None
    details: dict[int, dict] | None = None
    if args.known_accounts_only:
        print("Loading known accounts from CRM fxbackoffice.mt4_users ...")
        known_accounts, details = load_known_accounts_from_crm()
        print(f"  loaded {len(known_accounts):,} MT LOGINs with a real CRM userId")

    per_day, window, missing, account_servers = aggregate(
        dates, args.min_account_id, args.per_server, known_accounts
    )

    if missing:
        print(f"\n⚠️  Missing JSON for {len(missing)} day(s): {', '.join(missing)}")
    if not per_day:
        sys.exit("No data loaded — nothing to report.")

    # ------------------------------------------------------------------
    # Trade-activity enrichment (optional)
    # ------------------------------------------------------------------
    trade_counts: dict[int, int] | None = None
    trade_date: dt.date | None = None
    if args.enrich_trades:
        trade_date = parse_date_arg(args.enrich_trades)

        # Collect candidate accounts: only accounts on SHARED IPs (≥2 distinct
        # accounts on the same IP). No need to ping MySQL for singleton IPs.
        candidate_accounts: set[int] = set()
        for _bucket, ip_map in window.items():
            for _ip, v in ip_map.items():
                if len(v["accounts"]) >= 2:
                    candidate_accounts.update(v["accounts"])

        # Build the list of loginSid strings. account_servers tells us the
        # MT server(s) the account appeared on → maps to sid prefix.
        login_sid_to_account: dict[str, int] = {}
        for acc in candidate_accounts:
            for server in account_servers.get(acc, ()):
                prefix = SERVER_TO_SID_PREFIX.get(server)
                if prefix is None:
                    continue
                login_sid_to_account[f"{prefix}{acc}"] = acc

        print(
            f"\nEnriching with mt4_trades for {trade_date.isoformat()} — "
            f"checking {len(login_sid_to_account):,} loginSid(s) across "
            f"{len(candidate_accounts):,} shared-IP accounts ..."
        )
        raw_counts = load_trade_counts_from_crm(
            login_sid_to_account.keys(), trade_date
        )
        # Collapse back to per-account totals (MT4 + MT5 accounts sharing an
        # LOGIN would be summed, which is what a human reviewer wants).
        trade_counts = defaultdict(int)
        for login_sid, cnt in raw_counts.items():
            acc = login_sid_to_account.get(login_sid)
            if acc is not None:
                trade_counts[acc] += cnt
        trade_counts = dict(trade_counts)

        active = sum(1 for c in trade_counts.values() if c > 0)
        total = sum(trade_counts.values())
        print(
            f"  active accounts (traded on {trade_date.isoformat()}): {active:,} / "
            f"{len(candidate_accounts):,}   total orders: {total:,}"
        )

    print()
    print_per_day(per_day)
    print_window(window, len(dates))
    print_top_k(window, args.top_k, args.min_account_id)

    if args.csv:
        write_csv(window, Path(args.csv), details, trade_counts, trade_date)
    if args.csv_pivot:
        write_csv_pivot(
            window, Path(args.csv_pivot), details, trade_counts, trade_date
        )
    if args.xlsx:
        write_xlsx(window, Path(args.xlsx), details, trade_counts, trade_date)


if __name__ == "__main__":
    main()
